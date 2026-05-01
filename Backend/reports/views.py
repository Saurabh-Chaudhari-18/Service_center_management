"""
Reports views for business analytics and exports.

Features:
- Branch-wise revenue reports
- Pending jobs analysis
- Technician productivity
- Inventory consumption
- Export to Excel/PDF
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import models
from django.db.models import Sum, Count, Avg, F, Q
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from datetime import timedelta
from decimal import Decimal
import io

from core.permissions import CanViewReports
from core.models import Branch


class ReportsViewSet(viewsets.ViewSet):
    """
    ViewSet for generating various reports.
    Only Owners, Managers, and Accountants can view reports.
    """
    permission_classes = [IsAuthenticated, CanViewReports]

    def get_accessible_branches(self):
        """Get branches accessible to current user."""
        accessible = self.request.user.get_accessible_branches()
        branch_id = self.request.query_params.get('branch') or self.request.headers.get('X-Branch-ID')
        if branch_id:
            return accessible.filter(pk=branch_id)
        return accessible

    def get_date_range(self):
        """Parse date range from query params."""
        from_date = self.request.query_params.get('from_date')
        to_date = self.request.query_params.get('to_date')
        
        if not from_date:
            from_date = timezone.now().date() - timedelta(days=30)
        if not to_date:
            to_date = timezone.now().date()
        
        return from_date, to_date

    @action(detail=False, methods=['get'])
    def revenue(self, request):
        """
        Revenue report based on actual payments received.
        Filters by payment_date so "Today" shows payments collected today.
        """
        from billing.models import Payment, InvoiceStatus
        
        branches = self.get_accessible_branches()
        from_date, to_date = self.get_date_range()
        
        # Query payments by payment_date (when money was actually collected)
        payments = Payment.objects.filter(
            invoice__branch__in=branches,
            invoice__is_finalized=True,
            payment_date__date__gte=from_date,
            payment_date__date__lte=to_date,
            is_verified=True,
        ).exclude(invoice__status=InvoiceStatus.CANCELLED)
        
        # Summary by branch
        branch_summary = payments.values(
            'invoice__branch', 'invoice__branch__name'
        ).annotate(
            total_collected=Sum('amount'),
            payment_count=Count('id'),
        ).order_by('invoice__branch__name')
        
        # Rename keys for frontend compatibility
        branch_list = []
        for b in branch_summary:
            branch_list.append({
                'branch': b['invoice__branch'],
                'branch__name': b['invoice__branch__name'],
                'total_revenue': b['total_collected'],
                'invoice_count': b['payment_count'],
            })
        
        # Daily breakdown by payment date
        daily_revenue = payments.annotate(
            date=TruncDate('payment_date')
        ).values('date').annotate(
            revenue=Sum('amount'),
            collected=Sum('amount'),
            count=Count('id')
        ).order_by('date')
        
        # Calculate totals
        totals = payments.aggregate(
            total_revenue=Sum('amount'),
            total_collected=Sum('amount'),
            total_invoices=Count('invoice', distinct=True),
        )
        # Fill None values with 0
        totals = {k: v or 0 for k, v in totals.items()}
        totals['total_outstanding'] = 0
        totals['total_tax'] = 0
        
        # Recent payments detail (for owner visibility)
        recent_payments = payments.select_related(
            'invoice', 'received_by'
        ).order_by('-payment_date')[:20]
        
        payments_detail = [
            {
                'id': str(p.id),
                'payment_date': p.payment_date.isoformat(),
                'amount': float(p.amount),
                'payment_method': p.payment_method,
                'reference': p.reference,
                'invoice_number': p.invoice.invoice_number,
                'received_by': p.received_by.get_full_name() or p.received_by.email,
                'notes': p.notes,
            }
            for p in recent_payments
        ]
        
        return Response({
            'from_date': str(from_date),
            'to_date': str(to_date),
            'branches': branch_list,
            'daily_breakdown': list(daily_revenue),
            'totals': totals,
            'recent_payments': payments_detail,
        })

    @action(detail=False, methods=['get'])
    def pending_jobs(self, request):
        """
        Pending jobs analysis.
        Shows jobs by status, days pending, and branch.
        """
        from jobs.models import JobCard, JobStatus
        
        branches = self.get_accessible_branches()
        
        # Get all non-completed jobs
        pending_jobs = JobCard.objects.filter(
            branch__in=branches
        ).exclude(
            status__in=[JobStatus.DELIVERED, JobStatus.CANCELLED, JobStatus.REJECTED]
        )
        
        # Summary by status
        status_summary = pending_jobs.values('status').annotate(
            count=Count('id')
        ).order_by('status')
        
        # Summary by branch
        branch_summary = pending_jobs.values('branch', 'branch__name').annotate(
            count=Count('id'),
            urgent_count=Count('id', filter=Q(is_urgent=True))
        ).order_by('branch__name')
        
        # Overdue analysis (jobs pending more than expected)
        overdue_jobs = pending_jobs.filter(
            estimated_completion_date__lt=timezone.now().date()
        ).annotate(
            days_overdue=F('estimated_completion_date') - timezone.now().date()
        )
        
        # Age analysis
        today = timezone.now().date()
        age_groups = {
            '0-3 days': pending_jobs.filter(
                created_at__date__gte=today - timedelta(days=3)
            ).count(),
            '4-7 days': pending_jobs.filter(
                created_at__date__lt=today - timedelta(days=3),
                created_at__date__gte=today - timedelta(days=7)
            ).count(),
            '8-14 days': pending_jobs.filter(
                created_at__date__lt=today - timedelta(days=7),
                created_at__date__gte=today - timedelta(days=14)
            ).count(),
            '15+ days': pending_jobs.filter(
                created_at__date__lt=today - timedelta(days=14)
            ).count(),
        }
        
        return Response({
            'total_pending': pending_jobs.count(),
            'urgent_count': pending_jobs.filter(is_urgent=True).count(),
            'overdue_count': overdue_jobs.count(),
            'by_status': list(status_summary),
            'by_branch': list(branch_summary),
            'by_age': age_groups,
        })

    @action(detail=False, methods=['get'])
    def technician_productivity(self, request):
        """
        Technician productivity report.
        Shows jobs completed, average time, etc.
        """
        from jobs.models import JobCard, JobStatus
        from core.models import User, Role
        
        branches = self.get_accessible_branches()
        from_date, to_date = self.get_date_range()
        
        # Get technicians in accessible branches
        technicians = User.objects.filter(
            role=Role.TECHNICIAN,
            branches__in=branches,
            is_active=True
        ).distinct()
        
        productivity_data = []
        
        for tech in technicians:
            # Jobs assigned
            assigned_jobs = JobCard.objects.filter(
                assigned_technician=tech,
                branch__in=branches
            )
            
            # Completed in date range
            completed = assigned_jobs.filter(
                status=JobStatus.DELIVERED,
                delivery_date__date__gte=from_date,
                delivery_date__date__lte=to_date
            )
            
            # Currently assigned (not completed)
            current = assigned_jobs.exclude(
                status__in=[JobStatus.DELIVERED, JobStatus.CANCELLED, JobStatus.REJECTED]
            )
            
            productivity_data.append({
                'technician_id': str(tech.id),
                'technician_name': tech.get_full_name(),
                'jobs_completed': completed.count(),
                'jobs_in_progress': current.count(),
                'total_assigned': assigned_jobs.count(),
            })
        
        # Sort by jobs completed
        productivity_data.sort(key=lambda x: x['jobs_completed'], reverse=True)
        
        return Response({
            'from_date': str(from_date),
            'to_date': str(to_date),
            'technicians': productivity_data
        })

    @action(detail=False, methods=['get'])
    def inventory_consumption(self, request):
        """
        Inventory consumption report.
        Shows parts used over time period.
        """
        from inventory.models import JobPartUsage, InventoryItem
        
        branches = self.get_accessible_branches()
        from_date, to_date = self.get_date_range()
        
        # Get usage data
        usage = JobPartUsage.objects.filter(
            job__branch__in=branches,
            created_at__date__gte=from_date,
            created_at__date__lte=to_date
        )
        
        # Summary by item
        item_summary = usage.values(
            'inventory_item', 'inventory_item__name', 'inventory_item__sku'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_value=Sum('total_price'),
            usage_count=Count('id')
        ).order_by('-total_quantity')[:20]  # Top 20 items
        
        # Summary by category
        category_summary = usage.values(
            'inventory_item__category', 'inventory_item__category__name'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_value=Sum('total_price')
        ).order_by('-total_value')
        
        # Daily usage
        daily_usage = usage.annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            quantity=Sum('quantity'),
            value=Sum('total_price')
        ).order_by('date')
        
        # Totals
        totals = usage.aggregate(
            total_quantity=Sum('quantity'),
            total_value=Sum('total_price'),
            total_transactions=Count('id')
        )
        
        return Response({
            'from_date': str(from_date),
            'to_date': str(to_date),
            'top_items': list(item_summary),
            'by_category': list(category_summary),
            'daily_usage': list(daily_usage),
            'totals': totals
        })

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """
        Low stock report.
        Shows items below threshold.
        """
        from inventory.models import InventoryItem
        
        branches = self.get_accessible_branches()
        
        low_stock_items = InventoryItem.objects.filter(
            branch__in=branches,
            is_active=True,
            quantity__lte=F('low_stock_threshold')
        ).select_related('branch', 'category').order_by('quantity')
        
        data = []
        for item in low_stock_items:
            data.append({
                'id': str(item.id),
                'name': item.name,
                'sku': item.sku,
                'branch': item.branch.name,
                'category': item.category.name if item.category else None,
                'quantity': item.quantity,
                'threshold': item.low_stock_threshold,
                'shortage': max(0, item.low_stock_threshold - item.quantity),
                'cost_price': str(item.cost_price),
            })
        
        return Response({
            'total_items': len(data),
            'items': data
        })

    @action(detail=False, methods=['get'])
    def customer_analysis(self, request):
        """
        Customer analysis report.
        Shows top customers, repeat customers, etc.
        """
        from customers.models import Customer
        from billing.models import Invoice
        
        branches = self.get_accessible_branches()
        from_date, to_date = self.get_date_range()
        
        # Customers with invoices in period
        customers_with_revenue = Invoice.objects.filter(
            branch__in=branches,
            is_finalized=True,
            invoice_date__gte=from_date,
            invoice_date__lte=to_date
        ).values('job__customer', 'job__customer__first_name', 'job__customer__last_name', 'job__customer__mobile').annotate(
            total_revenue=Sum('total_amount'),
            invoice_count=Count('id')
        ).order_by('-total_revenue')[:20]
        
        # New customers in period
        new_customers = Customer.objects.filter(
            branch__in=branches,
            created_at__date__gte=from_date,
            created_at__date__lte=to_date
        ).count()
        
        # Total customers
        total_customers = Customer.objects.filter(
            branch__in=branches,
            is_active=True
        ).count()
        
        return Response({
            'from_date': str(from_date),
            'to_date': str(to_date),
            'total_customers': total_customers,
            'new_customers': new_customers,
            'top_customers': list(customers_with_revenue)
        })

    @action(detail=False, methods=['get'])
    def gst_summary(self, request):
        """
        GST summary report for filing.
        Shows CGST, SGST, IGST collected.
        """
        from billing.models import Invoice, InvoiceStatus
        
        branches = self.get_accessible_branches()
        from_date, to_date = self.get_date_range()
        
        invoices = Invoice.objects.filter(
            branch__in=branches,
            is_finalized=True,
            invoice_date__gte=from_date,
            invoice_date__lte=to_date
        ).exclude(status=InvoiceStatus.CANCELLED)
        
        # GST totals
        gst_summary = invoices.aggregate(
            total_taxable=Sum('subtotal'),
            total_cgst=Sum('cgst_total'),
            total_sgst=Sum('sgst_total'),
            total_igst=Sum('igst_total'),
            total_tax=Sum('total_tax'),
            total_value=Sum('total_amount'),
            invoice_count=Count('id')
        )
        
        # By GST rate
        from billing.models import InvoiceLineItem
        rate_summary = InvoiceLineItem.objects.filter(
            invoice__in=invoices
        ).values('gst_rate').annotate(
            taxable_amount=Sum('amount'),
            cgst_amount=Sum('cgst_amount'),
            sgst_amount=Sum('sgst_amount'),
            igst_amount=Sum('igst_amount'),
        ).order_by('gst_rate')
        
        # Intrastate vs Interstate
        supply_type = invoices.values('is_interstate').annotate(
            count=Count('id'),
            total=Sum('total_amount')
        )
        
        return Response({
            'from_date': str(from_date),
            'to_date': str(to_date),
            'summary': gst_summary,
            'by_rate': list(rate_summary),
            'by_supply_type': list(supply_type)
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export report data to Excel.
        """
        report_type = request.query_params.get('report', 'revenue')
        
        # Log export
        from audit.services import AuditLogService
        AuditLogService.log_export(
            user=request.user,
            export_type='EXCEL',
            report_name=report_type,
            parameters=dict(request.query_params)
        )
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            from django.http import HttpResponse
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report_type.capitalize()
            
            # Get report data based on type
            if report_type == 'revenue':
                data = self.revenue(request).data
                headers = ['Branch', 'Total Revenue', 'Collected', 'Outstanding', 'Invoices']
                ws.append(headers)
                for branch in data.get('branches', []):
                    ws.append([
                        branch.get('branch__name', ''),
                        branch.get('total_revenue', 0),
                        branch.get('total_collected', 0),
                        float(branch.get('total_revenue', 0) or 0) - float(branch.get('total_collected', 0) or 0),
                        branch.get('invoice_count', 0),
                    ])
            
            elif report_type == 'pending_jobs':
                data = self.pending_jobs(request).data
                headers = ['Status', 'Count']
                ws.append(headers)
                for status in data.get('by_status', []):
                    ws.append([status.get('status', ''), status.get('count', 0)])
            
            elif report_type == 'inventory':
                data = self.inventory_consumption(request).data
                headers = ['Item', 'SKU', 'Quantity Used', 'Total Value']
                ws.append(headers)
                for item in data.get('top_items', []):
                    ws.append([
                        item.get('inventory_item__name', ''),
                        item.get('inventory_item__sku', ''),
                        item.get('total_quantity', 0),
                        item.get('total_value', 0),
                    ])
            
            # Style header row
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = max_length + 2
            
            # Save to response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
            wb.save(response)
            
            return response
            
        except ImportError:
            return Response(
                {'error': 'Excel export requires openpyxl library.'},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )

    @action(detail=False, methods=['get'])
    def gstr1_export(self, request):
        """
        CA-Ready GSTR-1 format export.
        Exports finalized invoices in the standard GSTR-1 format with:
        - GSTIN, Invoice Number, Invoice Date, Invoice Value
        - Place of Supply, Taxable Value, CGST, SGST, IGST
        """
        from billing.models import Invoice, InvoiceStatus
        
        branches = self.get_accessible_branches()
        from_date, to_date = self.get_date_range()
        
        invoices = Invoice.objects.filter(
            branch__in=branches,
            is_finalized=True,
            invoice_date__gte=from_date,
            invoice_date__lte=to_date
        ).exclude(
            status=InvoiceStatus.CANCELLED
        ).select_related('branch').order_by('invoice_date')

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from django.http import HttpResponse

            wb = openpyxl.Workbook()

            # --- B2B Sheet (Business to Business) ---
            ws_b2b = wb.active
            ws_b2b.title = "B2B"
            b2b_headers = [
                'GSTIN of Recipient', 'Invoice Number', 'Invoice Date',
                'Invoice Value', 'Place of Supply', 'Reverse Charge',
                'Invoice Type', 'E-Commerce GSTIN', 'Rate',
                'Taxable Value', 'CGST Amount', 'SGST Amount', 'IGST Amount',
                'Cess Amount'
            ]
            ws_b2b.append(b2b_headers)

            # --- B2C Sheet (Business to Consumer) ---
            ws_b2c = wb.create_sheet("B2CL")
            b2c_headers = [
                'Invoice Number', 'Invoice Date', 'Invoice Value',
                'Place of Supply', 'Rate', 'Taxable Value',
                'CGST Amount', 'SGST Amount', 'IGST Amount', 'Cess Amount'
            ]
            ws_b2c.append(b2c_headers)

            for inv in invoices:
                state_code = inv.customer_state_code or (inv.branch.state_code if inv.branch else '')
                place_of_supply = f"{state_code}-{inv.branch.state if inv.branch else ''}"
                
                row_data = [
                    inv.invoice_number,
                    inv.invoice_date.strftime('%d-%m-%Y') if inv.invoice_date else '',
                    float(inv.total_amount),
                    place_of_supply,
                    float(inv.line_items.first().gst_rate if inv.line_items.exists() else 18),
                    float(inv.subtotal),
                    float(inv.cgst_total),
                    float(inv.sgst_total),
                    float(inv.igst_total),
                    0  # Cess
                ]

                if inv.customer_gstin:
                    # B2B
                    ws_b2b.append([
                        inv.customer_gstin,
                        *row_data[:3],
                        place_of_supply,
                        'N',  # Reverse Charge
                        'Regular',
                        '',  # E-Commerce GSTIN
                        *row_data[4:]
                    ])
                else:
                    # B2C
                    ws_b2c.append(row_data)

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for ws in [ws_b2b, ws_b2c]:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                # Auto-width
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max_len + 3

            # Log export
            from audit.services import AuditLogService
            AuditLogService.log_export(
                user=request.user,
                export_type='EXCEL',
                report_name='GSTR-1',
                parameters=dict(request.query_params)
            )

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="GSTR1_{from_date}_{to_date}.xlsx"'
            wb.save(response)
            return response

        except ImportError:
            return Response(
                {'error': 'Excel export requires openpyxl library.'},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )

    @action(detail=False, methods=['get'])
    def net_profit(self, request):
        """
        Net Profit calculation: Revenue - Expenses.
        Used for the dashboard's financial overview.
        """
        from billing.models import Payment, InvoiceStatus
        from expenses.models import Expense
        
        branches = self.get_accessible_branches()
        from_date, to_date = self.get_date_range()

        # Revenue = Sum of verified payments
        revenue = Payment.objects.filter(
            invoice__branch__in=branches,
            invoice__is_finalized=True,
            payment_date__date__gte=from_date,
            payment_date__date__lte=to_date,
            is_verified=True
        ).exclude(
            invoice__status=InvoiceStatus.CANCELLED
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        # Expenses = Sum of expense amounts
        expenses_total = Expense.objects.filter(
            branch__in=branches,
            expense_date__gte=from_date,
            expense_date__lte=to_date
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        net_profit = revenue - expenses_total

        return Response({
            'from_date': str(from_date),
            'to_date': str(to_date),
            'revenue': revenue,
            'expenses': expenses_total,
            'net_profit': net_profit,
            'profit_margin': round(float(net_profit) / float(revenue) * 100, 1) if revenue > 0 else 0
        })

