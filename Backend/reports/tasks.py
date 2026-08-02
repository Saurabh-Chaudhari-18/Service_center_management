"""
Celery tasks for async Excel report generation.

Flow:
  1. View dispatches task → returns {task_id, status_url}
  2. Client polls  GET /api/reports/export_status/?task_id=<id>
  3. When complete, status_url response contains {download_url}
  4. File is stored in media/exports/ and auto-cleaned after EXPORT_FILE_TTL seconds.
"""

import os
import uuid
import logging
from datetime import date, timedelta
from celery import shared_task
from django.conf import settings

logger = logging.getLogger(__name__)

EXPORT_DIR = os.path.join(settings.MEDIA_ROOT, 'exports')
EXPORT_FILE_TTL = 60 * 60  # 1 hour — Celery result backend cleans result; file cleaned manually


def _ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)


def _parse_dates(from_date_str, to_date_str):
    from datetime import datetime
    fmt = '%Y-%m-%d'
    from_date = datetime.strptime(from_date_str, fmt).date() if isinstance(from_date_str, str) else from_date_str
    to_date = datetime.strptime(to_date_str, fmt).date() if isinstance(to_date_str, str) else to_date_str
    return from_date, to_date


@shared_task(
    bind=True,
    max_retries=2,
    default_retry_delay=30,
    soft_time_limit=300,
    name='reports.generate_excel',
)
def generate_excel_report(self, report_type, branch_ids, from_date_str, to_date_str):
    """
    Generate an Excel report and save to media/exports/.
    Returns the relative media URL of the generated file.
    """
    import openpyxl
    from openpyxl.styles import Font
    from core.models import Branch
    from billing.models import Payment, Invoice, InvoiceStatus
    from jobs.models import JobCard

    _ensure_export_dir()
    from_date, to_date = _parse_dates(from_date_str, to_date_str)
    branches = Branch.objects.filter(id__in=branch_ids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.capitalize()

    if report_type == 'revenue':
        from django.db.models import Sum, Count
        rows = (
            Payment.objects
            .filter(invoice__branch__in=branches, payment_date__date__gte=from_date, payment_date__date__lte=to_date)
            .values('invoice__branch__name')
            .annotate(total_collected=Sum('amount'), invoice_count=Count('invoice', distinct=True))
        )
        headers = ['Branch', 'Total Collected', 'Invoice Count']
        ws.append(headers)
        for r in rows:
            ws.append([r['invoice__branch__name'], float(r['total_collected'] or 0), r['invoice_count']])

    elif report_type == 'pending_jobs':
        from django.db.models import Count
        rows = (
            JobCard.objects
            .filter(branch__in=branches)
            .exclude(status__in=['DELIVERED', 'CANCELLED', 'REJECTED'])
            .values('status')
            .annotate(count=Count('id'))
        )
        headers = ['Status', 'Count']
        ws.append(headers)
        for r in rows:
            ws.append([r['status'], r['count']])

    elif report_type == 'inventory':
        from inventory.models import JobPartUsage
        from django.db.models import Sum
        rows = (
            JobPartUsage.objects
            .filter(job__branch__in=branches, job__created_at__date__gte=from_date, job__created_at__date__lte=to_date)
            .values('inventory_item__name', 'inventory_item__sku')
            .annotate(total_quantity=Sum('quantity'), total_value=Sum('total_price'))
        )
        headers = ['Item', 'SKU', 'Quantity Used', 'Total Value']
        ws.append(headers)
        for r in rows:
            ws.append([
                r['inventory_item__name'], r['inventory_item__sku'],
                r['total_quantity'], float(r['total_value'] or 0),
            ])

    # Style header row
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 2

    filename = f"{report_type}_{from_date_str}_{to_date_str}_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return f"{settings.MEDIA_URL}exports/{filename}"


@shared_task(
    bind=True,
    max_retries=2,
    default_retry_delay=30,
    soft_time_limit=300,
    name='reports.generate_gstr1',
)
def generate_gstr1_report(self, branch_ids, from_date_str, to_date_str):
    """
    Generate CA-ready GSTR-1 Excel export and save to media/exports/.
    Returns the relative media URL of the generated file.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from core.models import Branch
    from billing.models import Invoice, InvoiceStatus

    _ensure_export_dir()
    from_date, to_date = _parse_dates(from_date_str, to_date_str)
    branches = Branch.objects.filter(id__in=branch_ids)

    invoices = (
        Invoice.objects
        .filter(branch__in=branches, is_finalized=True,
                invoice_date__gte=from_date, invoice_date__lte=to_date)
        .exclude(status=InvoiceStatus.CANCELLED)
        .select_related('branch')
        .prefetch_related('line_items')
        .order_by('invoice_date')
    )

    wb = openpyxl.Workbook()
    ws_b2b = wb.active
    ws_b2b.title = "B2B"
    b2b_headers = [
        'GSTIN of Recipient', 'Invoice Number', 'Invoice Date', 'Invoice Value',
        'Place of Supply', 'Reverse Charge', 'Invoice Type', 'E-Commerce GSTIN',
        'Rate', 'Taxable Value', 'CGST Amount', 'SGST Amount', 'IGST Amount', 'Cess Amount',
    ]
    ws_b2b.append(b2b_headers)

    ws_b2c = wb.create_sheet("B2CL")
    b2c_headers = [
        'Invoice Number', 'Invoice Date', 'Invoice Value', 'Place of Supply',
        'Rate', 'Taxable Value', 'CGST Amount', 'SGST Amount', 'IGST Amount', 'Cess Amount',
    ]
    ws_b2c.append(b2c_headers)

    for inv in invoices:
        from core.utils import format_place_of_supply
        state_code = inv.customer_state_code or (inv.branch.state_code if inv.branch else '')
        place_of_supply = format_place_of_supply(state_code)
        first_item = inv.line_items.first()
        gst_rate = float(first_item.gst_rate) if first_item else 18.0
        row_data = [
            inv.invoice_number,
            inv.invoice_date.strftime('%d-%m-%Y') if inv.invoice_date else '',
            float(inv.total_amount),
            place_of_supply, gst_rate, float(inv.subtotal),
            float(inv.cgst_total), float(inv.sgst_total), float(inv.igst_total), 0,
        ]
        if inv.customer_gstin:
            ws_b2b.append([inv.customer_gstin, *row_data[:3], place_of_supply, 'N', 'Regular', '', *row_data[4:]])
        else:
            ws_b2c.append(row_data)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for ws in [ws_b2b, ws_b2c]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 3

    filename = f"GSTR1_{from_date_str}_{to_date_str}_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return f"{settings.MEDIA_URL}exports/{filename}"
